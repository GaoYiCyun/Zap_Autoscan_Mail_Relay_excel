from zapv2 import ZAPv2
from datetime import datetime
from openpyxl import Workbook
from dotenv import load_dotenv
from email.message import EmailMessage
from collections import Counter

import os
import time
import smtplib

# === 載入環境變數 ===
load_dotenv()  # 若你放 email.env，沒放參數則會讀 .env

EMAIL_HOST = os.getenv('EMAIL_HOST')
EMAIL_PORT = int(os.getenv('EMAIL_PORT', '25'))   # 或者 587，視情況
EMAIL_FROM = os.getenv('EMAIL_FROM')
EMAIL_TO = os.getenv('EMAIL_TO_LIST')

if not EMAIL_TO:
    raise ValueError("EMAIL_TO_LIST 環境變數未設，請檢查 .env 檔案內容與位置。")
EMAIL_TO = [email.strip() for email in EMAIL_TO.split(',') if email.strip()]    # 支援多個信箱

# === ZAP 設定 ===
ZAP_API_KEY = '12345678'
ZAP_PROXY = 'http://127.0.0.1:8080'
TARGETS = [
    'http://192.168.149.129/DVWA/'
]
REPORT_DIR = 'zap_reports'

zap = ZAPv2(apikey=ZAP_API_KEY, proxies={'http': ZAP_PROXY, 'https': ZAP_PROXY})
os.makedirs(REPORT_DIR, exist_ok=True)

for target in TARGETS:
    now = datetime.now().strftime("%Y_%m_%d_%H_%M")
    
    # 建立 Excel 報表
    excel_file = f"{REPORT_DIR}/Risk_{target.replace('http://', '').replace('/', '_')}_{now}.xlsx"
    wb = Workbook()

    # 高風險工作表
    ws_high = wb.active
    ws_high.title = "High Risk Alerts"
    ws_high.append(["Alert", "Risk", "URL", "Description"])

    # 中風險工作表
    ws_medium = wb.create_sheet(title="Medium Risk Alerts")
    ws_medium.append(["Alert", "Risk", "URL", "Description"])

    # 低風險工作表
    ws_low = wb.create_sheet(title="Low Risk Alerts")
    ws_low.append(["Alert", "Risk", "URL", "Description"])

    # 警訊工作表
    ws_info = wb.create_sheet(title="Info Alerts")
    ws_info.append(["Alert", "Risk", "URL", "Description"])

    # 所有警報工作表
    ws_all = wb.create_sheet(title="All Alerts")
    ws_all.append(["Alert", "Risk", "URL", "Description"])

    print(f"\n📍 開始掃描：{target}")

    print("🔍 Spidering...")
    scan_id = zap.spider.scan(target)
    while int(zap.spider.status(scan_id)) < 100:
        print(f"  Spider Progress: {zap.spider.status(scan_id)}%")
        time.sleep(2)

    print("💥 Active Scanning...")
    scan_id = zap.ascan.scan(target)
    while int(zap.ascan.status(scan_id)) < 100:
        print(f"  Active Scan Progress: {zap.ascan.status(scan_id)}%")
        time.sleep(5)

    print("📄 收集漏洞...")
    
    alerts = zap.core.alerts()
    from collections import Counter
    risks = [a['risk'].strip().lower() for a in alerts]
    risk_counter = Counter(risks)
    print("[DEBUG] 所有風險統計：", Counter(risks))
    
    for alert in alerts:
        alert_name = alert['alert']
        risk = alert['risk'].strip().lower()
        url = alert['url']
        desc = alert.get('description', '')[:100]

        # 統一 row 資料
        row = [alert_name, risk, url, desc]

        # 寫入 all 分頁
        ws_all.append(row)

        # 高風險
        if risk == 'high':
            ws_high.append(row)
        # 中風險
        elif risk == 'medium':
            ws_medium.append(row)
        elif risk == 'low':
            ws_low.append(row)
        elif risk == 'informational':
            ws_info.append(row)

    wb.save(excel_file)

    print("📄 輸出 HTML 報告...")
    domain = target.replace("http://", "").replace("https://", "").replace("/", "_")
    html_file = f"{REPORT_DIR}/{domain}_{now}.html"
    with open(html_file, "w", encoding="utf-8") as f:
        f.write(zap.core.htmlreport())

    def send_email_to_list(to_list, subject, body, attachments):
        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = 'Zap_report'
        msg['To'] = ', '.join(to_list)
        msg.set_content(body)
        for filepath in attachments:
            if os.path.exists(filepath):
                with open(filepath, 'rb') as f:
                    file_data = f.read()
                    file_name = os.path.basename(filepath)
                    msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)
            else:
                print(f"附件檔案不存在: {filepath}")

        with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT, timeout=10) as smtp:
            smtp.set_debuglevel(1)
            smtp.ehlo()
            smtp.send_message(msg)
            print("📧 郵件已寄出")
    # 寄送剛剛生成的報告
    levels = ['high', 'medium', 'low', 'informational']
    level_names = {
        'high': '高風險',
        'medium': '中風險',
        'low': '低風險',
        'informational': '資訊'
    }
    body_lines = []
    for lv in levels:
        count = risk_counter.get(lv, 0)
        body_lines.append(f"{level_names[lv]}：{count} 件")

    mail_body = (
        f"{now} ZAP 掃描漏洞統計如下：\n\n"
        + '\n'.join(body_lines)
        + "\n\n詳情請見附件報告。"
    )
    send_email_to_list(
        to_list=EMAIL_TO,
        subject=f'ZAP 掃描報告 - {now}',
        body=mail_body,
        attachments=[html_file, excel_file]
    )

print("\n✅ 所有目標掃描完成，報告已寄出！")